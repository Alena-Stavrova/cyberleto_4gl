from openpyxl import load_workbook
import math

def round_price(value):
    return math.floor(value + 0.5)

def read_promo_excel(filepath):
    workbook = load_workbook(filepath, data_only=True)
    sheet = workbook.active  

    headers = [cell.value for cell in sheet[1]]
    sku_col = headers.index("Артикул")
    name_col = headers.index("Номенклатура")
    old_price_col = headers.index("Розница")
    discount_col= headers.index("Скидка")
    new_price_col = headers.index("Розница со скидкой")


    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        items.append({
            'sku': str(row[sku_col]),
            'name': row[name_col],
            'old_price': str(row[old_price_col]),
            'discount' : str(row[discount_col]),
            'new_price': str(round_price(row[new_price_col]))
        })

    return items

if __name__ == "__main__":
    items = read_promo_excel("item_list.xlsx")
    print(f"Read {len(items)} items")
    print(items[0])   # First item
    print(items[-1])  # Last item
