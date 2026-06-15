from openpyxl import load_workbook
import math

def round_price(value):
    return math.floor(value + 0.5)

workbook = load_workbook("item_list.xlsx", data_only=True)
sheet = workbook.active  

# sheet[1] is row 1. Each cell has a .value attribute.
headers = [cell.value for cell in sheet[1]]
sku_col = headers.index("Артикул")
name_col = headers.index("Наименование")
old_price_col = headers.index("Текущая цена, руб.")
discount_col= headers.index("Скидка")
new_price_col = headers.index("Конечная цена")


items = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    items.append({
        'sku': str(row[sku_col]),
        'name': row[name_col],
        'old_price': str(row[old_price_col]),
        'discount' : str(row[discount_col]),
        'new_price': str(round_price(row[new_price_col]))
        #'new_price' : str(extract_price(row[new_price_col]))
    })
print(items)
print(len(items))
print(items[0]['discount'], items[1]['discount'], items[2]['discount'])

